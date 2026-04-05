"""
Build the Phase F ATO Codes reference workbook.
Sheets (per user's requested structure):
  1. Current Structure          — current categories/subcategories from src/types/financials.ts
  2. ATO Personal Structure     — D-codes + income items from the Individual Tax Return
  3. Mapping: Current -> Personal  — each current sub mapped to a personal ATO code (or none)
  4. ATO Business Structure     — Item 6 codes from the Company Tax Return (Pty Ltd)
  5. Mapping: Current -> Business  — each current sub mapped to a company ATO code (or none)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).resolve().parents[1] / "docs" / "reference" / "phase-f-ato-codes.xlsx"

wb = Workbook()

HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
SECTION_FONT = Font(bold=True, color="1E3A8A", size=10)
ALT_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 22

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def style_body(ws, start_row, ncols, zebra=True):
    last = ws.max_row
    for r in range(start_row, last + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
            if zebra and (r - start_row) % 2 == 1:
                # only apply alt fill if no section fill is already set
                if cell.fill.start_color.rgb in (None, "00000000"):
                    cell.fill = ALT_FILL

# ============================================================
# Sheet 1 — Current Structure (as-is, from src/types/financials.ts)
# ============================================================
ws1 = wb.active
ws1.title = "1. Current Structure"

CATEGORIES = {
    "INCOME": ["Salary / Payroll", "Freelance / Consulting", "Investment Returns", "Government Benefits", "Other Income"],
    "BUSINESS EXPENSES": ["Contractors & Services", "Equipment & Technology", "Professional Services", "Marketing & Advertising"],
    "DINING & COFFEE": ["Lunch", "Dinner Out", "Takeaway", "Cafes & Coffee", "Drinks / Bar"],
    "DONATIONS & GIVING": ["Church / Religious", "Charity", "Community / School", "Sponsorship"],
    "EDUCATION & CHILDCARE": ["School Fees", "Childcare", "Courses & Books"],
    "FINANCIAL": ["Bank Fees", "Credit Card Payments", "Loan Repayments"],
    "GROCERIES": ["Supermarket", "Butcher / Bakery", "Market / Deli"],
    "HEALTH": ["Medical / GP / Specialist", "Pharmacy", "Dental", "Allied Health"],
    "HOUSING": ["Mortgage / Rent", "Strata / Body Corporate", "Maintenance & Repairs"],
    "HOUSEHOLD BILLS": ["Electricity", "Gas", "Water", "Council Rates", "Internet", "Phone / Mobile"],
    "INSURANCE": ["Life & Income Protection", "Health Insurance", "Home & Contents", "Car / Motor Vehicle", "Pet Insurance", "Business Insurance"],
    "INVESTMENTS": ["Shares / ETFs", "Managed Funds", "Super Contributions", "Property Investment", "Crypto", "Dividends & Distributions"],
    "KIDS & PETS": ["Kids Activities", "Kids Clothing", "Pet Food", "Vet", "Pet Grooming", "Dog Walker / Daycare"],
    "SHOPPING & LIFESTYLE": ["Clothing & Apparel", "Electronics & Tech", "Home & Garden", "Gifts"],
    "SPORTS & FITNESS": ["Soccer / Football", "Gym & Fitness", "Swimming", "Sports Equipment", "Other Sports"],
    "SUBSCRIPTIONS & DIGITAL": ["Streaming", "Software & SaaS", "Cloud Services", "Telco", "News & Media"],
    "TRANSFERS": ["Internal Transfer", "Family Transfer"],
    "TRANSPORT": ["Tesla / EV Charging", "Fuel", "Rego & CTP", "Tolls & Parking", "Public Transport", "Uber / Rideshare"],
    "TRAVEL": ["Flights", "Hotels & Accommodation", "Car Rental", "Travel Insurance", "Activities & Tours"],
    "OTHER": [],
}

ws1.append(["#", "Category", "Subcategory"])
style_header(ws1, 3)
n = 0
for cat, subs in CATEGORIES.items():
    if not subs:
        n += 1
        ws1.append([n, cat, "(none)"])
    else:
        for s in subs:
            n += 1
            ws1.append([n, cat, s])
autosize(ws1, [6, 26, 36])
style_body(ws1, 2, 3)
ws1.freeze_panes = "A2"

# ============================================================
# Sheet 2 — ATO Personal Structure (Individual Tax Return)
# ============================================================
ws2 = wb.create_sheet("2. ATO Personal")
ws2.append(["Code", "Form label", "Use"])
style_header(ws2, 3)

personal = [
    ("— Income —", "", ""),
    ("I-1",  "Item 1 — Salary, wages, allowances",       "PAYG income"),
    ("I-10", "Item 10 — Gross interest",                 "Bank interest received"),
    ("I-11", "Item 11 — Dividends",                      "Franked / unfranked dividends"),
    ("I-13", "Item 13 — Partnerships & trusts",          "Trust / partnership distributions"),
    ("I-18", "Item 18 — Capital gains",                  "Realised CGT events"),
    ("I-24", "Item 24 — Other income",                   "Freelance, gov benefits, misc"),
    ("— Deductions —", "", ""),
    ("D1",  "D1 — Work-related car expenses",            "Car for work (cents/km or logbook)"),
    ("D2",  "D2 — Work-related travel expenses",         "Flights, hotels for work (non-car)"),
    ("D3",  "D3 — Work-related clothing/laundry",        "Uniforms, protective, laundry"),
    ("D4",  "D4 — Self-education expenses",              "Courses tied to current role"),
    ("D5",  "D5 — Other work-related expenses",          "Phone %, WFH %, tools"),
    ("D9",  "D9 — Gifts or donations",                   "DGR-registered charities only"),
    ("D10", "D10 — Cost of managing tax affairs",        "Accountant, tax software"),
    ("D12", "D12 — Personal super contributions",        "Concessional voluntary contributions"),
    ("D15", "D15 — Other deductions",                    "Income protection premiums, etc."),
]
for row in personal:
    ws2.append(list(row))
    if row[0].startswith("—"):
        for c in range(1, 4):
            cell = ws2.cell(row=ws2.max_row, column=c)
            cell.fill = SECTION_FILL
            cell.font = SECTION_FONT
autosize(ws2, [10, 40, 46])
style_body(ws2, 2, 3)
ws2.freeze_panes = "A2"

# ============================================================
# Sheet 3 — Mapping Current -> Personal ATO
# ============================================================
ws3 = wb.create_sheet("3. Map Current → Personal")
ws3.append(["Current category", "Current subcategory", "Personal ATO code", "Notes"])
style_header(ws3, 4)

# "" = non-deductible / not applicable for personal return
map_personal = [
    ("INCOME", "Salary / Payroll",              "I-1",   ""),
    ("INCOME", "Freelance / Consulting",        "I-24",  "If not through a Pty Ltd"),
    ("INCOME", "Investment Returns",            "I-10 / I-11", "Split by source: interest → I-10, dividends → I-11"),
    ("INCOME", "Government Benefits",           "I-24",  ""),
    ("INCOME", "Other Income",                  "I-24",  ""),
    ("BUSINESS EXPENSES", "Contractors & Services",   "",   "Only relevant for company entities (see sheet 5)"),
    ("BUSINESS EXPENSES", "Equipment & Technology",   "",   "Only relevant for company entities"),
    ("BUSINESS EXPENSES", "Professional Services",    "D10","D10 if tax/accounting fees; else company expense"),
    ("BUSINESS EXPENSES", "Marketing & Advertising",  "",   "Only relevant for company entities"),
    ("DINING & COFFEE", "(all)",                "",      "Private — non-deductible"),
    ("DONATIONS & GIVING", "Church / Religious","D9 (if DGR)", "Most churches are not DGRs — verify per merchant"),
    ("DONATIONS & GIVING", "Charity",           "D9 (if DGR)", "DGR register check required"),
    ("DONATIONS & GIVING", "Community / School","D9 (if DGR)", ""),
    ("DONATIONS & GIVING", "Sponsorship",       "D9 (if DGR)", "Usually non-deductible"),
    ("EDUCATION & CHILDCARE", "School Fees",    "",      "Non-deductible"),
    ("EDUCATION & CHILDCARE", "Childcare",      "",      "Non-deductible (offset elsewhere)"),
    ("EDUCATION & CHILDCARE", "Courses & Books","D4",    "Only if tied to current role"),
    ("FINANCIAL", "Bank Fees",                  "",      "Non-deductible on personal accounts"),
    ("FINANCIAL", "Credit Card Payments",       "",      "Transfer — excluded from reports"),
    ("FINANCIAL", "Loan Repayments",            "D5 / D7", "Interest portion only, if investment-linked; principal not deductible"),
    ("GROCERIES", "(all)",                      "",      "Non-deductible"),
    ("HEALTH", "(all)",                         "",      "Non-deductible (private health → separate offset)"),
    ("HOUSING", "Mortgage / Rent",              "",      "Home residence — not deductible"),
    ("HOUSING", "Strata / Body Corporate",      "",      "Home residence — not deductible"),
    ("HOUSING", "Maintenance & Repairs",        "",      "Home residence — not deductible"),
    ("HOUSEHOLD BILLS", "Electricity",          "D5",    "× WFH % from assumptions register"),
    ("HOUSEHOLD BILLS", "Gas",                  "D5",    "× WFH %"),
    ("HOUSEHOLD BILLS", "Water",                "",      "Non-deductible"),
    ("HOUSEHOLD BILLS", "Council Rates",        "",      "Non-deductible"),
    ("HOUSEHOLD BILLS", "Internet",             "D5",    "× work-use %"),
    ("HOUSEHOLD BILLS", "Phone / Mobile",       "D5",    "× work-use %"),
    ("INSURANCE", "Life & Income Protection",   "D15",   "Income-protection portion only"),
    ("INSURANCE", "Health Insurance",           "",      "Offset — not a deduction"),
    ("INSURANCE", "Home & Contents",            "",      "Non-deductible"),
    ("INSURANCE", "Car / Motor Vehicle",        "",      "Non-deductible (personal)"),
    ("INSURANCE", "Pet Insurance",              "",      "Non-deductible"),
    ("INSURANCE", "Business Insurance",         "",      "Only for company entities"),
    ("INVESTMENTS", "Shares / ETFs",            "",      "Capital — CGT on disposal (out of scope)"),
    ("INVESTMENTS", "Managed Funds",            "",      "Capital — CGT on disposal"),
    ("INVESTMENTS", "Super Contributions",      "D12",   "Concessional voluntary only"),
    ("INVESTMENTS", "Property Investment",      "",      "Capital — rental schedule separate (out of scope)"),
    ("INVESTMENTS", "Crypto",                   "",      "Capital — CGT on disposal"),
    ("INVESTMENTS", "Dividends & Distributions","I-11",  "Income, not a deduction"),
    ("KIDS & PETS", "(all)",                    "",      "Non-deductible"),
    ("SHOPPING & LIFESTYLE", "Clothing & Apparel","D3",  "Only compulsory uniforms / PPE"),
    ("SHOPPING & LIFESTYLE", "Electronics & Tech","D5",  "× work-use % for personal; depreciation if capital"),
    ("SHOPPING & LIFESTYLE", "Home & Garden",   "",      "Non-deductible"),
    ("SHOPPING & LIFESTYLE", "Gifts",           "",      "Non-deductible"),
    ("SPORTS & FITNESS", "(all)",               "",      "Non-deductible"),
    ("SUBSCRIPTIONS & DIGITAL", "Streaming",    "",      "Non-deductible"),
    ("SUBSCRIPTIONS & DIGITAL", "Software & SaaS","D5",  "× work-use %"),
    ("SUBSCRIPTIONS & DIGITAL", "Cloud Services","D5",   "× work-use %"),
    ("SUBSCRIPTIONS & DIGITAL", "Telco",        "D5",    "× work-use %"),
    ("SUBSCRIPTIONS & DIGITAL", "News & Media", "",      "Non-deductible unless job-required"),
    ("TRANSFERS", "(all)",                      "",      "Auto-excluded from reports"),
    ("TRANSPORT", "Tesla / EV Charging",        "D1",    "× work-use %"),
    ("TRANSPORT", "Fuel",                       "D1",    "× work-use %"),
    ("TRANSPORT", "Rego & CTP",                 "D1",    "× work-use %"),
    ("TRANSPORT", "Tolls & Parking",            "D1",    "× work-use %"),
    ("TRANSPORT", "Public Transport",           "D2",    "Only when travelling for work"),
    ("TRANSPORT", "Uber / Rideshare",           "D2",    "Only when travelling for work"),
    ("TRAVEL", "Flights",                       "D2",    "Work travel only"),
    ("TRAVEL", "Hotels & Accommodation",        "D2",    "Work travel only"),
    ("TRAVEL", "Car Rental",                    "D2",    "Work travel only"),
    ("TRAVEL", "Travel Insurance",              "",      "Non-deductible"),
    ("TRAVEL", "Activities & Tours",            "",      "Non-deductible"),
    ("OTHER", "(empty)",                        "",      ""),
]
for row in map_personal:
    ws3.append(list(row))
autosize(ws3, [22, 34, 22, 48])
style_body(ws3, 2, 4)
ws3.freeze_panes = "A2"

# ============================================================
# Sheet 4 — ATO Business Structure (Company Tax Return, Item 6)
# ============================================================
ws4 = wb.create_sheet("4. ATO Business (Pty Ltd)")
ws4.append(["Code", "Form label", "Use"])
style_header(ws4, 3)

company = [
    ("— Income —", "", ""),
    ("6-INCOME",    "Item 6 — Total income",                     "Gross sales / services revenue"),
    ("6-INT-REC",   "Item 6 — Gross interest (received)",        "Interest earned on business accounts"),
    ("6-DIV-REC",   "Item 6 — Total dividends (received)",       "Dividend income"),
    ("6-OTHER-INC", "Item 6 — Other gross income",               "Misc business income"),
    ("— Direct expenses (own line on Item 6) —", "", ""),
    ("6-COGS",      "Item 6 — Cost of sales",                    "Direct costs / purchases / stock movement"),
    ("6-CONTRACT",  "Item 6 — Contractor / sub-contractor / commission", "ABN-holder payments"),
    ("6-WAGES",     "Item 6 — Total salary and wage expenses",   "Employee PAYG wages"),
    ("6-SUPER",     "Item 6 — Superannuation expenses",          "Employer SG contributions"),
    ("6-BAD-DEBT",  "Item 6 — Bad debts",                        "Written-off receivables"),
    ("6-LEASE",     "Item 6 — Lease expenses (Australia)",       "Operating leases"),
    ("6-RENT",      "Item 6 — Rent expenses",                    "Premises rent"),
    ("6-INT-PAID",  "Item 6 — Interest expenses (Australia)",    "Business loan interest"),
    ("6-DEPN",      "Item 6 — Depreciation expenses",            "Capital asset writedown"),
    ("6-MV",        "Item 6 — Motor vehicle expenses",           "Business vehicle running costs"),
    ("6-REPAIRS",   "Item 6 — Repairs and maintenance",          "Non-capital fixes"),
    ("6-FBT",       "Item 6 — Fringe benefits tax",              "FBT paid"),
    ("6-AUDIT",     "Item 6 — External audit fees",              "If audited"),
    ("— Catchall line on the form —", "", ""),
    ("6-OTHER-EXP", "Item 6 — All other expenses",               "Form rollup line for everything below"),
    ("— Internal sub-codes (roll up to 6-OTHER-EXP on the form) —", "", ""),
    ("6-OTHER-INSURANCE", "(internal) Business insurance",        "PI, public liability, contents"),
    ("6-OTHER-MARKETING", "(internal) Marketing & advertising",   "Ads, promo, content"),
    ("6-OTHER-SUBS",      "(internal) Subscriptions & SaaS",      "Software tools"),
    ("6-OTHER-TELCO",     "(internal) Telco & internet",          "Business phone, data"),
    ("6-OTHER-PROFEES",   "(internal) Professional fees",         "Legal, accounting, consulting"),
    ("6-OTHER-BANKFEES",  "(internal) Bank & merchant fees",      "Incl. Stripe/Square fees"),
    ("6-OTHER-OFFICE",    "(internal) Office supplies",           "Consumables, stationery"),
    ("6-OTHER-TRAVEL",    "(internal) Business travel (non-MV)",  "Flights, accommodation"),
    ("6-OTHER-MISC",      "(internal) Other / misc",              "Catchall"),
]
for row in company:
    ws4.append(list(row))
    if row[0].startswith("—"):
        for c in range(1, 4):
            cell = ws4.cell(row=ws4.max_row, column=c)
            cell.fill = SECTION_FILL
            cell.font = SECTION_FONT
autosize(ws4, [22, 48, 44])
style_body(ws4, 2, 3)
ws4.freeze_panes = "A2"

# ============================================================
# Sheet 5 — Mapping Current -> Business ATO (bonus sheet)
# ============================================================
ws5 = wb.create_sheet("5. Map Current → Business")
ws5.append(["Current category", "Current subcategory", "Business ATO code", "Notes"])
style_header(ws5, 4)

# "" = not applicable to a company context
map_business = [
    ("INCOME", "Salary / Payroll",              "",           "Personal income — not a company line"),
    ("INCOME", "Freelance / Consulting",        "6-INCOME",   "If billed through a Pty Ltd"),
    ("INCOME", "Investment Returns",            "6-INT-REC / 6-DIV-REC", "Split by source"),
    ("INCOME", "Government Benefits",           "",           "Personal only"),
    ("INCOME", "Other Income",                  "6-OTHER-INC",""),
    ("BUSINESS EXPENSES", "Contractors & Services", "6-CONTRACT", ""),
    ("BUSINESS EXPENSES", "Equipment & Technology", "6-DEPN",  "Depreciation if > $300, else 6-OTHER-OFFICE"),
    ("BUSINESS EXPENSES", "Professional Services", "6-OTHER-PROFEES", "Legal, accounting, consulting"),
    ("BUSINESS EXPENSES", "Marketing & Advertising","6-OTHER-MARKETING",""),
    ("DINING & COFFEE", "(all)",                "",           "Non-deductible (entertainment)"),
    ("DONATIONS & GIVING", "(all)",             "",           "Companies use different deduction rules (out of scope)"),
    ("EDUCATION & CHILDCARE", "Courses & Books","6-OTHER-EXP","Staff training only"),
    ("FINANCIAL", "Bank Fees",                  "6-OTHER-BANKFEES", ""),
    ("FINANCIAL", "Credit Card Payments",       "",           "Transfer — excluded"),
    ("FINANCIAL", "Loan Repayments",            "6-INT-PAID", "Interest portion only; principal not a P&L item"),
    ("GROCERIES", "(all)",                      "",           "Not a company expense"),
    ("HEALTH", "(all)",                         "",           "Not a company expense"),
    ("HOUSING", "Mortgage / Rent",              "6-RENT",     "Only if business premises"),
    ("HOUSING", "Strata / Body Corporate",      "6-RENT",     "Business premises only"),
    ("HOUSING", "Maintenance & Repairs",        "6-REPAIRS",  "Business premises only"),
    ("HOUSEHOLD BILLS", "Electricity",          "6-OTHER-OFFICE", "If business premises or home office"),
    ("HOUSEHOLD BILLS", "Gas",                  "6-OTHER-OFFICE", "If business premises"),
    ("HOUSEHOLD BILLS", "Water",                "6-OTHER-OFFICE", "If business premises"),
    ("HOUSEHOLD BILLS", "Council Rates",        "6-OTHER-OFFICE", "If business premises"),
    ("HOUSEHOLD BILLS", "Internet",             "6-OTHER-TELCO", ""),
    ("HOUSEHOLD BILLS", "Phone / Mobile",       "6-OTHER-TELCO", ""),
    ("INSURANCE", "Life & Income Protection",   "",           "Personal only"),
    ("INSURANCE", "Business Insurance",         "6-OTHER-INSURANCE", ""),
    ("INSURANCE", "Other personal insurance",   "",           "Not a company line"),
    ("INVESTMENTS", "Super Contributions",      "6-SUPER",    "Employer SG paid by company"),
    ("INVESTMENTS", "Other investments",        "",           "CGT tracked separately"),
    ("KIDS & PETS", "(all)",                    "",           "Not a company expense"),
    ("SHOPPING & LIFESTYLE", "Clothing & Apparel","6-OTHER-EXP", "PPE / uniforms for staff"),
    ("SHOPPING & LIFESTYLE", "Electronics & Tech","6-DEPN",    "Depreciation if > $300"),
    ("SHOPPING & LIFESTYLE", "Home & Garden",   "",           "Not a company expense"),
    ("SHOPPING & LIFESTYLE", "Gifts",           "",           "FBT implications — generally disallowed"),
    ("SPORTS & FITNESS", "(all)",               "",           "Not a company expense"),
    ("SUBSCRIPTIONS & DIGITAL", "Streaming",    "",           "Not a company expense"),
    ("SUBSCRIPTIONS & DIGITAL", "Software & SaaS","6-OTHER-SUBS",""),
    ("SUBSCRIPTIONS & DIGITAL", "Cloud Services","6-OTHER-SUBS",""),
    ("SUBSCRIPTIONS & DIGITAL", "Telco",        "6-OTHER-TELCO",""),
    ("SUBSCRIPTIONS & DIGITAL", "News & Media", "6-OTHER-MISC","If trade publications"),
    ("TRANSFERS", "(all)",                      "",           "Auto-excluded"),
    ("TRANSPORT", "Tesla / EV Charging",        "6-MV",       "If company vehicle"),
    ("TRANSPORT", "Fuel",                       "6-MV",       "If company vehicle"),
    ("TRANSPORT", "Rego & CTP",                 "6-MV",       "If company vehicle"),
    ("TRANSPORT", "Tolls & Parking",            "6-MV",       "If company vehicle"),
    ("TRANSPORT", "Public Transport",           "6-OTHER-TRAVEL", "Staff work travel"),
    ("TRANSPORT", "Uber / Rideshare",           "6-OTHER-TRAVEL", "Staff work travel"),
    ("TRAVEL", "Flights",                       "6-OTHER-TRAVEL", "Business travel"),
    ("TRAVEL", "Hotels & Accommodation",        "6-OTHER-TRAVEL", "Business travel"),
    ("TRAVEL", "Car Rental",                    "6-OTHER-TRAVEL", "Business travel"),
    ("TRAVEL", "Travel Insurance",              "6-OTHER-TRAVEL", "Business travel"),
    ("TRAVEL", "Activities & Tours",            "",           "Entertainment — generally disallowed"),
    ("OTHER", "(empty)",                        "",           ""),
]
for row in map_business:
    ws5.append(list(row))
autosize(ws5, [22, 34, 26, 48])
style_body(ws5, 2, 4)
ws5.freeze_panes = "A2"

# ============================================================
# Save
# ============================================================
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Wrote {OUT}")
